#!/usr/bin/env python3
"""
build_liquiditaetsplan.py
=========================

Erstellt eine rollierende Liquiditätsvorschau (13 / 26 / 52 Wochen) im
Layout der Klotzkette-Vorlage `vorlage-liquiditaetsplan.xlsx`:

    Bestand Start Planung
      Kassenbestand
      Kontostand
      Liquidität Wochenanfang            (hellblau, fett)
    Einnahmen                            (gelb, fett)
      Umsatzerlöse
      erhaltene Anzahlungen
      sonstige Einnahmen
      Summe Einnahmen
     Ausgaben                            (rot, fett)
      Löhne und Gehälter einschl. LSt
      Sozialversicherung
      Waren/ Material
      Miete
      Heizung/Strom/Wasser
      Verwaltung
      Werbung/Marketing
      Fahrzeug-/Leasingkosten
      Versicherungen
      Beratungskosten
      sonstige Ausgaben
      Investitionen
      Betriebliche Steuern
      sonstige Ausgaben (II)
      Darlehenstilgung
      Zinsen
      Privatentnahmen
      Summe Ausgaben
    Cash Flow Woche
    Liquidität Woche Ende                (hellblau, fett)

Erweitert um:
    fällige Verbindlichkeiten Folgewoche
    3-Wochen-Lücke (kumuliert)
    Lücken-Quote (%)
    Ampel § 17 InsO  (bedingte Formatierung GRÜN/GELB/ROT)

Sheets: 13-Wochen, 26-Wochen, 52-Wochen, Fortfuehrungsprognose, Annahmen.

Eingabe: YAML- oder JSON-Datei. Beispiel-YAML siehe Beispielakte.

Verwendung:
    python build_liquiditaetsplan.py --eingabe mandant.yaml \\
                                     --ausgabe liquiditaetsplan.xlsx

Quellen:
- BGH, Urt. v. 24.05.2005 – IX ZR 123/04, BGHZ 163, 134 Rn. 12 ff. (§ 17 InsO)
- BGH, Urt. v. 19.11.2019 – II ZR 233/18, NJW 2020, 1809 Rn. 17 ff.
- IDW S 6 (Sanierungskonzepte), IDW S 11 (Insolvenzeröffnungsgründe)
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    import yaml  # type: ignore
except ImportError:
    yaml = None


# ----------------------------------------------------------------------------
# Eingabe
# ----------------------------------------------------------------------------

def load_input(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8-sig")
    if path.suffix.lower() in {".yaml", ".yml"}:
        if yaml is None:
            return load_simple_yaml(text, path)
        return yaml.safe_load(text) or {}
    if path.suffix.lower() == ".json":
        return json.loads(text)
    raise SystemExit(f"Unbekanntes Eingabeformat: {path.suffix}")


def load_simple_yaml(text: str, path: Path) -> dict[str, Any]:
    """Kleiner YAML-Fallback für die mitgelieferten Akten, falls PyYAML fehlt."""
    root: dict[Any, Any] = {}
    stack: list[tuple[int, dict[Any, Any] | list[Any]]] = [(-1, root)]
    lines: list[tuple[int, int, str]] = []

    for lineno, raw_line in enumerate(text.splitlines(), start=1):
        if "\t" in raw_line[: len(raw_line) - len(raw_line.lstrip())]:
            raise SystemExit(f"{path}:{lineno}: Tabs in Einrückungen werden nicht unterstützt.")

        line = strip_yaml_comment(raw_line).rstrip()
        if not line.strip():
            continue

        indent = len(line) - len(line.lstrip(" "))
        content = line[indent:]
        lines.append((lineno, indent, content))

    for index, (lineno, indent, content) in enumerate(lines):
        while indent <= stack[-1][0]:
            stack.pop()
        parent = stack[-1][1]

        if content.startswith("- "):
            if not isinstance(parent, list):
                raise SystemExit(f"{path}:{lineno}: YAML-Liste ohne Listenkontext.")

            item_text = content[2:].strip()
            split_at = find_yaml_key_separator(item_text)
            if item_text and split_at is not None:
                item: dict[Any, Any] = {}
                parent.append(item)
                stack.append((indent, item))

                key_text = item_text[:split_at].strip()
                value_text = item_text[split_at + 1:].strip()
                if not key_text:
                    raise SystemExit(f"{path}:{lineno}: Leerer YAML-Schlüssel.")
                key = parse_yaml_scalar(key_text)
                if value_text == "":
                    child: dict[Any, Any] | list[Any]
                    child = [] if next_yaml_child_is_list(lines, index, indent) else {}
                    item[key] = child
                    stack.append((next_yaml_child_stack_indent(lines, index, indent), child))
                else:
                    item[key] = parse_yaml_scalar(value_text)
            elif item_text:
                parent.append(parse_yaml_scalar(item_text))
            else:
                item = {}
                parent.append(item)
                stack.append((next_yaml_child_stack_indent(lines, index, indent), item))
            continue

        if not isinstance(parent, dict):
            raise SystemExit(f"{path}:{lineno}: YAML-Schlüssel ohne Mapping-Kontext.")

        split_at = find_yaml_key_separator(content)
        if split_at is None:
            raise SystemExit(f"{path}:{lineno}: YAML-Zeile nicht verstanden: {content!r}")

        key_text = content[:split_at].strip()
        value_text = content[split_at + 1:].strip()
        if not key_text:
            raise SystemExit(f"{path}:{lineno}: Leerer YAML-Schlüssel.")

        key = parse_yaml_scalar(key_text)

        if value_text == "":
            child: dict[Any, Any] | list[Any]
            child = [] if next_yaml_child_is_list(lines, index, indent) else {}
            parent[key] = child
            stack.append((next_yaml_child_stack_indent(lines, index, indent), child))
        else:
            parent[key] = parse_yaml_scalar(value_text)

    return root


def next_yaml_child_is_list(lines: list[tuple[int, int, str]], index: int, indent: int) -> bool:
    for _, next_indent, content in lines[index + 1:]:
        if next_indent <= indent:
            return False
        if next_indent > indent:
            return content.startswith("- ")
    return False


def next_yaml_child_stack_indent(lines: list[tuple[int, int, str]], index: int, indent: int) -> int:
    for _, next_indent, _ in lines[index + 1:]:
        if next_indent > indent:
            return next_indent - 1
    return indent


def strip_yaml_comment(line: str) -> str:
    quote: str | None = None
    escaped = False
    for i, ch in enumerate(line):
        if quote == '"':
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == quote:
                quote = None
        elif quote == "'":
            if ch == quote:
                quote = None
        else:
            if ch in {"'", '"'}:
                quote = ch
            elif ch == "#":
                return line[:i]
    return line


def find_yaml_key_separator(content: str) -> int | None:
    quote: str | None = None
    escaped = False
    for i, ch in enumerate(content):
        if quote == '"':
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == quote:
                quote = None
        elif quote == "'":
            if ch == quote:
                quote = None
        else:
            if ch in {"'", '"'}:
                quote = ch
            elif ch == ":":
                return i
    return None


def parse_yaml_scalar(value: str) -> Any:
    if value.startswith('"') and value.endswith('"'):
        return json.loads(value)
    if value.startswith("'") and value.endswith("'"):
        return value[1:-1].replace("''", "'")

    lower = value.lower()
    if lower in {"null", "~"}:
        return None
    if lower == "true":
        return True
    if lower == "false":
        return False

    normalized = value.replace("_", "")
    try:
        return int(normalized)
    except ValueError:
        pass
    try:
        return float(normalized)
    except ValueError:
        return value


# ----------------------------------------------------------------------------
# Styles (1:1 nach Vorlage)
# ----------------------------------------------------------------------------

FONT_DEFAULT = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", size=10, bold=True)

FILL_HEADER_GRAY = PatternFill("solid", fgColor="D8D8D8")
FILL_LIQUI_BLUE = PatternFill("solid", fgColor="00CCFF")
FILL_EIN_YELLOW = PatternFill("solid", fgColor="FFFF00")
FILL_AUS_RED = PatternFill("solid", fgColor="FF0000")
FILL_SUM_LIGHT = PatternFill("solid", fgColor="F2F2F2")

NUM_FMT = '#,##0_);[Red](#,##0)'
NUM_FMT_PCT = '0.0%'

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ----------------------------------------------------------------------------
# Wochenraster
# ----------------------------------------------------------------------------

def monday_of(d: dt.date) -> dt.date:
    return d - dt.timedelta(days=d.weekday())


def week_buckets(start: dt.date, n_weeks: int) -> list[tuple[int, dt.date, dt.date]]:
    mon = monday_of(start)
    weeks = []
    for i in range(n_weeks):
        a = mon + dt.timedelta(days=7 * i)
        b = a + dt.timedelta(days=6)
        _, iso_week, _ = a.isocalendar()
        weeks.append((iso_week, a, b))
    return weeks


# Vorlagen-Layout: Spalte A = Label, Spalte B = "Bestand Start"-Spalte (KW vor Start),
# Spalten C ff. = Planungs-KW.
LABEL_COL = 1   # A
START_COL = 2   # B = Bestand Start (KW vor Stichtag)
PLAN_COL_START = 3   # C ff. = KW 1 ff.

# Zeilen exakt wie Vorlage:
R_TITLE = 1
R_STAND = 2
R_BESTAND_HEADER = 6
R_KASSE = 7
R_KONTO = 8
R_LIQUI_START = 9
R_EIN_HEADER = 11
R_EIN_UMSATZ = 12
R_EIN_ANZAHLUNG = 13
R_EIN_SONST = 14
R_EIN_SUMME = 15
R_AUS_HEADER = 17
R_AUS_LOHN = 18
R_AUS_SV = 19
R_AUS_WAREN = 20
R_AUS_MIETE = 21
R_AUS_ENERGIE = 22
R_AUS_VERWALTUNG = 23
R_AUS_WERBUNG = 24
R_AUS_LEASING = 25
R_AUS_VERSICH = 26
R_AUS_BERATUNG = 27
R_AUS_SONST = 28
R_AUS_INVEST = 29
R_AUS_STEUERN = 30
R_AUS_SONST2 = 31
R_AUS_TILGUNG = 32
R_AUS_ZINSEN = 33
R_AUS_ENTNAHME = 34
R_AUS_SUMME = 35
R_CASHFLOW = 37
R_LIQUI_END = 39
# Klotzkette-Ergänzung: § 17 InsO-Block
R_FAELLIG = 41
R_LUECKE_3W = 42
R_QUOTE = 43
R_AMPEL = 44

# Mapping: Bucket-Key -> Zeile
EIN_KEYS = {
    "umsatz": R_EIN_UMSATZ,
    "anzahlungen": R_EIN_ANZAHLUNG,
    "sonstige": R_EIN_SONST,
}
AUS_KEYS = {
    "lohn": R_AUS_LOHN,
    "sv": R_AUS_SV,
    "waren": R_AUS_WAREN,
    "miete": R_AUS_MIETE,
    "energie": R_AUS_ENERGIE,
    "verwaltung": R_AUS_VERWALTUNG,
    "werbung": R_AUS_WERBUNG,
    "leasing": R_AUS_LEASING,
    "versicherung": R_AUS_VERSICH,
    "beratung": R_AUS_BERATUNG,
    "sonstige": R_AUS_SONST,
    "investitionen": R_AUS_INVEST,
    "steuern": R_AUS_STEUERN,
    "sonstige2": R_AUS_SONST2,
    "tilgung": R_AUS_TILGUNG,
    "zinsen": R_AUS_ZINSEN,
    "entnahmen": R_AUS_ENTNAHME,
}


def build_skeleton(ws, daten: dict, n_weeks: int):
    """Statische Zellen und Beschriftungen anlegen."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions[get_column_letter(START_COL)].width = 12
    for i in range(n_weeks):
        ws.column_dimensions[get_column_letter(PLAN_COL_START + i)].width = 13

    # Titel
    c = ws.cell(row=R_TITLE, column=1, value="Liquiditätsplan")
    c.font = FONT_BOLD
    ws.cell(row=R_STAND, column=1,
            value=f"Stand: {daten.get('stichtag', '')} – {daten.get('mandant', {}).get('name', '')}").font = FONT_DEFAULT

    # Header Zeile 6 – Bestand Start
    h = ws.cell(row=R_BESTAND_HEADER, column=LABEL_COL, value="Bestand Start Planung")
    h.font = FONT_BOLD
    h.fill = FILL_HEADER_GRAY
    h.number_format = NUM_FMT

    # KW-Header
    _, iso_week, _ = monday_of(dt.date.fromisoformat(daten["stichtag"])).isocalendar()
    prev_kw = iso_week - 1 if iso_week > 1 else 52
    kw_cell = ws.cell(row=R_BESTAND_HEADER, column=START_COL, value=f"KW {prev_kw}")
    kw_cell.font = FONT_BOLD
    kw_cell.alignment = Alignment(horizontal="center")

    stichtag = dt.date.fromisoformat(daten["stichtag"])
    for i, (kw, a, b) in enumerate(week_buckets(stichtag, n_weeks)):
        c = ws.cell(row=R_BESTAND_HEADER, column=PLAN_COL_START + i, value=f"KW {kw}")
        c.font = FONT_BOLD
        c.alignment = Alignment(horizontal="center")

    # Zeilen-Labels
    def put_label(row, text, bold=False, fill=None, numfmt=None):
        cell = ws.cell(row=row, column=LABEL_COL, value=text)
        if bold:
            cell.font = FONT_BOLD
        else:
            cell.font = FONT_DEFAULT
        if fill is not None:
            cell.fill = fill
        if numfmt:
            cell.number_format = numfmt

    put_label(R_KASSE, "Kassenbestand")
    put_label(R_KONTO, "Kontostand")
    put_label(R_LIQUI_START, "Liquidität Wochenanfang", bold=True, fill=FILL_LIQUI_BLUE, numfmt=NUM_FMT)

    put_label(R_EIN_HEADER, "Einnahmen", bold=True, fill=FILL_EIN_YELLOW, numfmt=NUM_FMT)
    put_label(R_EIN_UMSATZ, "Umsatzerlöse")
    put_label(R_EIN_ANZAHLUNG, "erhaltene Anzahlungen")
    put_label(R_EIN_SONST, "sonstige Einnahmen")
    put_label(R_EIN_SUMME, "Summe Einnahmen", bold=True, numfmt=NUM_FMT)

    put_label(R_AUS_HEADER, " Ausgaben", bold=True, fill=FILL_AUS_RED, numfmt=NUM_FMT)
    put_label(R_AUS_LOHN, "Löhne und Gehälter einschl. LSt")
    put_label(R_AUS_SV, "Sozialversicherung")
    put_label(R_AUS_WAREN, "Waren/ Material")
    put_label(R_AUS_MIETE, "Miete")
    put_label(R_AUS_ENERGIE, "Heizung/Strom/Wasser")
    put_label(R_AUS_VERWALTUNG, "Verwaltung")
    put_label(R_AUS_WERBUNG, "Werbung/Marketing")
    put_label(R_AUS_LEASING, "Fahrzeug-/Leasingkosten")
    put_label(R_AUS_VERSICH, "Versicherungen")
    put_label(R_AUS_BERATUNG, "Beratungskosten")
    put_label(R_AUS_SONST, "sonstige Ausgaben")
    put_label(R_AUS_INVEST, "Investitionen")
    put_label(R_AUS_STEUERN, "Betriebliche Steuern")
    put_label(R_AUS_SONST2, "sonstige Ausgaben")
    put_label(R_AUS_TILGUNG, "Darlehenstilgung")
    put_label(R_AUS_ZINSEN, "Zinsen")
    put_label(R_AUS_ENTNAHME, "Privatentnahmen")
    put_label(R_AUS_SUMME, "Summe Ausgaben", bold=True, numfmt=NUM_FMT)

    put_label(R_CASHFLOW, "Cash Flow Woche", bold=True, numfmt=NUM_FMT)
    put_label(R_LIQUI_END, "Liquidität Woche Ende", bold=True, fill=FILL_LIQUI_BLUE, numfmt=NUM_FMT)

    # Ampel-Block (Klotzkette-Erweiterung)
    put_label(R_FAELLIG, "fällige Verb. Folgewoche", bold=True, numfmt=NUM_FMT)
    put_label(R_LUECKE_3W, "3-Wochen-Lücke (kumuliert)", bold=True, numfmt=NUM_FMT)
    put_label(R_QUOTE, "Lücken-Quote (%)", bold=True, numfmt=NUM_FMT_PCT)
    put_label(R_AMPEL, "Ampel § 17 InsO", bold=True)


def fill_sheet(ws, daten: dict, n_weeks: int):
    """Werte und Formeln in die Spalten B (Start) und C..(C+n) eintragen."""
    # Bestand Start (Spalte B = KW vor Stichtag)
    kasse_start = float(daten.get("kassenbestand_start", 0))
    konto_start = float(daten.get("kontostand_start", 0))
    ws.cell(row=R_KASSE, column=START_COL, value=kasse_start).number_format = NUM_FMT
    ws.cell(row=R_KONTO, column=START_COL, value=konto_start).number_format = NUM_FMT
    ws.cell(row=R_LIQUI_START, column=START_COL,
            value=f"={get_column_letter(START_COL)}{R_KASSE}+{get_column_letter(START_COL)}{R_KONTO}"
            ).number_format = NUM_FMT
    ws.cell(row=R_LIQUI_START, column=START_COL).font = FONT_BOLD
    ws.cell(row=R_LIQUI_START, column=START_COL).fill = FILL_LIQUI_BLUE

    plan = daten.get("plan", {}) or {}

    for i in range(n_weeks):
        col = PLAN_COL_START + i
        L = get_column_letter(col)
        L_prev = get_column_letter(col - 1)
        kw_data: dict[str, Any] = {}
        if isinstance(plan, dict):
            # Plan-Keys koennen int (YAML) oder str (JSON) sein.
            kw_data = plan.get(i, plan.get(str(i), {}))
        elif isinstance(plan, list) and i < len(plan):
            kw_data = plan[i]

        # Liquidität Wochenanfang = Liq. Woche Ende der Vorwoche
        # (für Spalte C = i=0 → B (Bestand Start), sonst Vorspalte R_LIQUI_END)
        if i == 0:
            ws.cell(row=R_LIQUI_START, column=col,
                    value=f"={get_column_letter(START_COL)}{R_LIQUI_START}")
        else:
            ws.cell(row=R_LIQUI_START, column=col,
                    value=f"={L_prev}{R_LIQUI_END}")
        ws.cell(row=R_LIQUI_START, column=col).font = FONT_BOLD
        ws.cell(row=R_LIQUI_START, column=col).fill = FILL_LIQUI_BLUE
        ws.cell(row=R_LIQUI_START, column=col).number_format = NUM_FMT

        # Einnahmen
        ein = kw_data.get("einnahmen", {}) if isinstance(kw_data, dict) else {}
        for key, row in EIN_KEYS.items():
            val = float(ein.get(key, 0))
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = NUM_FMT
        # Summe Einnahmen
        c = ws.cell(row=R_EIN_SUMME, column=col,
                    value=f"=SUM({L}{R_EIN_UMSATZ}:{L}{R_EIN_SONST})")
        c.font = FONT_BOLD
        c.number_format = NUM_FMT

        # Ausgaben
        aus = kw_data.get("ausgaben", {}) if isinstance(kw_data, dict) else {}
        for key, row in AUS_KEYS.items():
            val = float(aus.get(key, 0))
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = NUM_FMT
        # Summe Ausgaben
        c = ws.cell(row=R_AUS_SUMME, column=col,
                    value=f"=SUM({L}{R_AUS_LOHN}:{L}{R_AUS_ENTNAHME})")
        c.font = FONT_BOLD
        c.number_format = NUM_FMT

        # Cash Flow Woche
        c = ws.cell(row=R_CASHFLOW, column=col,
                    value=f"={L}{R_EIN_SUMME}-{L}{R_AUS_SUMME}")
        c.font = FONT_BOLD
        c.number_format = NUM_FMT

        # Liquidität Woche Ende
        c = ws.cell(row=R_LIQUI_END, column=col,
                    value=f"={L}{R_LIQUI_START}+{L}{R_CASHFLOW}")
        c.font = FONT_BOLD
        c.fill = FILL_LIQUI_BLUE
        c.number_format = NUM_FMT

        # § 17 InsO-Block: fällige Verbindlichkeiten Folgewoche
        faellig = float(kw_data.get("faellig_folgewoche", 0))
        c = ws.cell(row=R_FAELLIG, column=col, value=faellig)
        c.number_format = NUM_FMT

    # 3-Wochen-Lücke und Quote/Ampel – muss in zweitem Durchgang, da Bezug auf Nachbarspalten
    for i in range(n_weeks):
        col = PLAN_COL_START + i
        L = get_column_letter(col)
        # Lücke = MAX(0, Summe fälliger Verb. t..t+2 - (Liquidität Woche Ende_t + Summe Einnahmen_t+1 + Summe Einnahmen_t+2))
        # Quote = Lücke / Summe fälliger Verb. t..t+2 (gleiche Bezugsgröße wie Lücke — BGH IX ZR 171/15)
        if i + 2 < n_weeks:
            L1 = get_column_letter(col + 1)
            L2 = get_column_letter(col + 2)
            faellig_sum = f"{L}{R_FAELLIG}+{L1}{R_FAELLIG}+{L2}{R_FAELLIG}"
            mittel = f"{L}{R_LIQUI_END}+{L1}{R_EIN_SUMME}+{L2}{R_EIN_SUMME}"
            ws.cell(row=R_LUECKE_3W, column=col,
                    value=f"=MAX(0,({faellig_sum})-({mittel}))").number_format = NUM_FMT
            ws.cell(row=R_QUOTE, column=col,
                    value=f"=IFERROR({L}{R_LUECKE_3W}/({faellig_sum}),0)"
                    ).number_format = NUM_FMT_PCT
        else:
            ws.cell(row=R_LUECKE_3W, column=col,
                    value=f"=MAX(0,{L}{R_FAELLIG}-{L}{R_LIQUI_END})").number_format = NUM_FMT
            ws.cell(row=R_QUOTE, column=col,
                    value=f"=IFERROR({L}{R_LUECKE_3W}/{L}{R_FAELLIG},0)"
                    ).number_format = NUM_FMT_PCT
        # Ampel
        c = ws.cell(row=R_AMPEL, column=col,
                    value=(f'=IF({L}{R_QUOTE}>=0.1,"ROT",'
                           f'IF({L}{R_QUOTE}>0,"GELB","GRÜN"))'))
        c.font = FONT_BOLD
        c.alignment = Alignment(horizontal="center")

    # bedingte Formatierung Ampel
    first_col = get_column_letter(PLAN_COL_START)
    last_col = get_column_letter(PLAN_COL_START + n_weeks - 1)
    rng = f"{first_col}{R_AMPEL}:{last_col}{R_AMPEL}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"ROT"'],
                        fill=PatternFill("solid", fgColor="F4B7BA"),
                        font=Font(bold=True, color="9C0006")))
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"GELB"'],
                        fill=PatternFill("solid", fgColor="FFE9A8"),
                        font=Font(bold=True, color="8F6A00")))
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"GRÜN"'],
                        fill=PatternFill("solid", fgColor="C9E5C1"),
                        font=Font(bold=True, color="2C6E1F")))

    ws.freeze_panes = ws.cell(row=R_BESTAND_HEADER + 1, column=PLAN_COL_START)


# ----------------------------------------------------------------------------
# Fortführungsprognose und Annahmen
# ----------------------------------------------------------------------------

def build_prognose_sheet(wb, daten):
    ws = wb.create_sheet("Fortfuehrungsprognose")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 22

    t = ws.cell(row=1, column=1, value="Fortführungsprognose nach IDW S 6 / IDW S 11")
    t.font = Font(name="Calibri", size=14, bold=True)

    for i, h in enumerate(["IDW-S-6-Element", "Befund / Annahme", "Status"]):
        c = ws.cell(row=3, column=1 + i, value=h)
        c.font = FONT_BOLD
        c.fill = FILL_HEADER_GRAY
        c.alignment = Alignment(horizontal="center")

    elemente = [
        ("1. Auftrag/Auftragsdurchführung", daten.get("prognose", {}).get("auftrag", "")),
        ("2. Basisinformationen", daten.get("prognose", {}).get("basis", "")),
        ("3. Krisenstadium und Ursachen", daten.get("prognose", {}).get("krise", "")),
        ("4. Leitbild des sanierten Unternehmens", daten.get("prognose", {}).get("leitbild", "")),
        ("5. Maßnahmen zur Krisenbewältigung", daten.get("prognose", {}).get("massnahmen", "")),
        ("6. Integrierte Planung 24+ Monate", daten.get("prognose", {}).get("planung", "")),
        ("7. Aussage Sanierungsfähigkeit", daten.get("prognose", {}).get("sanierung", "")),
        ("8. Fortbestehensprognose § 19 InsO", daten.get("prognose", {}).get("fortbestehen", "")),
    ]
    for i, (label, befund) in enumerate(elemente):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = FONT_BOLD
        ws.cell(row=r, column=2, value=befund).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=3, value="").alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 60
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = BORDER

    r = 4 + len(elemente) + 2
    ws.cell(row=r, column=1, value="Quellen:").font = FONT_BOLD
    quellen = [
        "BGH, Urt. v. 24.05.2005 – IX ZR 123/04, BGHZ 163, 134 Rn. 12 ff. (§ 17 InsO).",
        "BGH, Urt. v. 12.10.2006 – IX ZR 228/03, NJW 2007, 78 Rn. 16 ff. (Indizienkatalog).",
        "BGH, Urt. v. 19.11.2019 – II ZR 233/18, NJW 2020, 1809 Rn. 17 ff. (Fortbestehensprognose).",
        "BGH, Urt. v. 09.10.2012 – II ZR 298/11, BGHZ 195, 42 Rn. 12 ff. (insolvenzrechtl. Überschuldung).",
        "IDW S 6 (Sanierungskonzepte), IDW S 11 (Insolvenzeröffnungsgründe).",
        "K. Schmidt/Herchen, K. Schmidt, InsO, 20. Aufl. 2023, §§ 17, 19 InsO.",
        "Quellenregel: Keine Kommentar-, Handbuch-, Aufsatz- oder Tabellenfundstellen aus Modellwissen; nur Nutzerquelle, amtliche/freie Quelle oder lizenzierte Live-Verifikation verwenden.",
    ]
    for i, q in enumerate(quellen):
        ws.cell(row=r + 1 + i, column=1, value="• " + q)


def build_annahmen_sheet(wb, daten):
    ws = wb.create_sheet("Annahmen")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 60

    ws.cell(row=1, column=1, value="Annahmen und Inputs").font = Font(name="Calibri", size=14, bold=True)
    rows = [
        ("Mandant", daten.get("mandant", {}).get("name", "")),
        ("Rechtsform", daten.get("mandant", {}).get("rechtsform", "")),
        ("Branche", daten.get("mandant", {}).get("branche", "")),
        ("Mitarbeiter", daten.get("mandant", {}).get("mitarbeiter", "")),
        ("Stichtag", daten.get("stichtag", "")),
        ("Kassenbestand Start", daten.get("kassenbestand_start", 0)),
        ("Kontostand Start", daten.get("kontostand_start", 0)),
        ("Kreditlinie gesamt", daten.get("kreditlinie_gesamt", 0)),
        ("Kreditlinie genutzt", daten.get("kreditlinie_genutzt", 0)),
        ("Szenario", daten.get("szenario", "Base")),
    ]
    for i, (k, v) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=k).font = FONT_BOLD
        ws.cell(row=r, column=2, value=v)


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--eingabe", "-i", required=True, type=Path)
    parser.add_argument("--ausgabe", "-o", required=True, type=Path)
    args = parser.parse_args()

    daten = load_input(args.eingabe)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, n in [("13-Wochen", 13), ("26-Wochen", 26), ("52-Wochen", 52)]:
        ws = wb.create_sheet(name)
        build_skeleton(ws, daten, n)
        fill_sheet(ws, daten, n)

    build_prognose_sheet(wb, daten)
    build_annahmen_sheet(wb, daten)

    wb.save(args.ausgabe)
    print(f"OK: geschrieben nach {args.ausgabe}")


if __name__ == "__main__":
    main()
